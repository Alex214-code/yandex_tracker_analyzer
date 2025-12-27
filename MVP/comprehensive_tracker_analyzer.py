#!/usr/bin/env python3
"""
Комплексный анализатор данных Яндекс Трекера
Адаптирован под требования:
- Помесячная разбивка (одна строка = одна задача в одном месяце)
- Учет статусов на начало месяца
- Учет переходов статусов внутри месяца
- Логика "Активного пула" (был в работе на начало или перешел в работу)
- Разделы определяются по корневой родительской задаче (Контейнеру)
- Подсчет уровня вложенности
"""

import os
import time
import random
import argparse
from datetime import datetime, timedelta
from typing import Dict, List, Optional, Tuple
from dataclasses import dataclass
from pathlib import Path
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from loguru import logger
from dotenv import load_dotenv

# Загружаем переменные окружения
load_dotenv()

# Настройка логирования
logger.remove()
logger.add("tracker_analyzer.log", rotation="1 day", retention="7 days", level="INFO", 
           format="{time:YYYY-MM-DD HH:mm:ss} | {level} | {message}")
logger.add(lambda msg: print(msg, end=""), level="INFO", format="{message}")

# Конфигурация API
CLIENT_ID = os.getenv("YANDEX_CLIENT_ID")
OAUTH_TOKEN = os.getenv("YANDEX_OAUTH_TOKEN")
X_ORG_ID = os.getenv("YANDEX_ORG_ID")

if not all([CLIENT_ID, OAUTH_TOKEN, X_ORG_ID]):
    logger.error("ОШИБКА: Не найдены обязательные настройки в .env файле.")
    logger.error("Пожалуйста, проверьте YANDEX_CLIENT_ID, YANDEX_OAUTH_TOKEN и YANDEX_ORG_ID.")
    exit(1)

API_BASE_URL = "https://api.tracker.yandex.net/v2"
HEADERS = {
    "Authorization": f"OAuth {OAUTH_TOKEN}",
    "X-Org-ID": X_ORG_ID,
    "Content-Type": "application/json"
}

class TaskStatus:
    OPEN = "open"
    IN_PROGRESS = "inProgress" 
    PAUSED = "paused"
    CLOSED = "closed"
    NEED_INFO = "needInfo"

    ALL_STATUSES = [OPEN, IN_PROGRESS, PAUSED, CLOSED, NEED_INFO]
    
    @classmethod
    def get_display_name(cls, status_key: str) -> str:
        mapping = {
            cls.OPEN: "Открыт",
            cls.IN_PROGRESS: "В работе",
            cls.PAUSED: "Приостановлено",
            cls.CLOSED: "Закрыт",
            cls.NEED_INFO: "Требуется информация",
            "onHold": "Приостановлено"
        }
        return mapping.get(status_key, status_key)

TARGET_PROJECTS = [
    "НорНикель НОФ (Тех. Поддержка)",
    "УГМК Святогор (Тех. Поддержка)",
    "УГМК Гайский ГОК (Тех. Поддержка)",
    "РМК Томинский ГОК (Тех. Поддержка)"
]

@dataclass
class TaskChange:
    timestamp: datetime
    field: str
    old_value: str
    new_value: str

@dataclass
class Task:
    key: str
    summary: str
    project: str
    assignee: str
    status: str
    created: datetime
    updated: datetime
    resolved: Optional[datetime] = None
    parent_key: Optional[str] = None  # Ключ родительской задачи
    changes: List[TaskChange] = None
    priority: str = ""
    
    def __post_init__(self):
        if self.changes is None:
            self.changes = []

class YandexTrackerAnalyzer:
    
    def __init__(self):
        self.session = requests.Session()
        self.session.headers.update(HEADERS)
        self.changelog_cache = {}
        self.task_cache = {} # Кэш задач для построения иерархии
        
    def get_task_details(self, task_key: str) -> Optional[Dict]:
        """Получает информацию о задаче (с кэшированием)."""
        if task_key in self.task_cache:
            return self.task_cache[task_key]
            
        try:
            # Запрашиваем только нужные поля для оптимизации
            response = self.session.get(
                f"{API_BASE_URL}/issues/{task_key}", 
                params={"fields": "summary,parent"}, 
                timeout=30
            )
            if response.status_code == 404:
                return None
            response.raise_for_status()
            data = response.json()
            self.task_cache[task_key] = data
            return data
        except Exception as e:
            logger.warning(f"Не удалось получить информацию о задаче {task_key}: {e}")
            return None

    def resolve_root_section(self, task: Task) -> Tuple[str, int]:
        """
        Возвращает (Название раздела, Уровень вложенности).
        Раздел - это Summary корневой задачи (у которой нет родителя).
        Если у исходной задачи сразу нет родителя -> Раздел = "Контейнер".
        """
        depth = 0
        current_key = task.key
        
        # Если у задачи сразу нет родителя
        if not task.parent_key:
            return "Контейнер", 0

        # Идем вверх по цепочке
        parent_key = task.parent_key
        root_summary = "Не определен"
        
        while parent_key:
            depth += 1
            parent_data = self.get_task_details(parent_key)
            
            if not parent_data:
                # Если родитель не найден или нет доступа, прерываемся
                break
                
            root_summary = parent_data.get('summary', 'Без названия')
            
            # Проверяем следующего родителя
            parent_obj = parent_data.get('parent')
            if parent_obj and isinstance(parent_obj, dict) and 'key' in parent_obj:
                current_key = parent_key
                parent_key = parent_obj['key']
            else:
                # Дошли до верха
                break
                
        return root_summary, depth

    def fetch_changelogs_concurrently(self, tasks: List[Dict]):
        tasks_to_fetch = [t['key'] for t in tasks if t['key'] not in self.changelog_cache]
        if not tasks_to_fetch:
            return

        logger.info(f"Загружаю историю изменений для {len(tasks_to_fetch)} задач...")
        
        def fetch_one(task_key):
            url = f"{API_BASE_URL}/issues/{task_key}/changelog"
            retries = 3
            for i in range(retries):
                try:
                    response = self.session.get(url, timeout=30)
                    if response.status_code == 429:
                        time.sleep(2 + random.random())
                        continue
                    response.raise_for_status()
                    return task_key, response.json()
                except:
                    if i == retries - 1:
                        return task_key, []
                    time.sleep(1)
            return task_key, []

        with ThreadPoolExecutor(max_workers=5) as executor:
            future_to_key = {executor.submit(fetch_one, key): key for key in tasks_to_fetch}
            for i, future in enumerate(as_completed(future_to_key)):
                key, changelog = future.result()
                self.changelog_cache[key] = changelog
                if (i + 1) % 50 == 0:
                    logger.info(f"  ...загружено {i + 1} историй")

    def get_tasks_by_project(self, project_name: str, year: int, month: int) -> List[Dict]:
        tasks = []
        page = 1
        
        first_day = datetime(year, month, 1)
        if month == 12:
            next_month = datetime(year + 1, 1, 1)
        else:
            next_month = datetime(year, month + 1, 1)
        
        date_from = first_day.strftime('%Y-%m-%d')
        date_to = next_month.strftime('%Y-%m-%d')
        
        logger.info(f"Запрашиваю задачи проекта '{project_name}' за {month:02d}.{year}")
        
        filter_query = (
            f'Project: "{project_name}" AND ('
            f'(Updated: >= "{date_from}" AND Updated: < "{date_to}") OR '
            f'(Created: < "{date_to}" AND (Resolved: empty() OR Resolved: >= "{date_from}"))'
            f')'
        )
        
        while True:
            try:
                response = self.session.post(
                    f"{API_BASE_URL}/issues/_search",
                    params={"perPage": 100, "page": page},
                    json={"query": filter_query},
                    timeout=30
                )
                if response.status_code == 429:
                    time.sleep(5)
                    continue
                response.raise_for_status()
                data = response.json()
                if not data:
                    break
                tasks.extend(data)
                page += 1
            except Exception as e:
                logger.error(f"Ошибка при получении задач: {e}")
                break
                
        return tasks
    
    def get_task_changelog(self, task_key: str) -> List[Dict]:
        if task_key in self.changelog_cache:
            return self.changelog_cache[task_key]
        try:
            response = self.session.get(f"{API_BASE_URL}/issues/{task_key}/changelog", timeout=30)
            response.raise_for_status()
            return response.json()
        except:
            return []
    
    def extract_status_changes(self, task_key: str, changelog: List[Dict]) -> List[TaskChange]:
        changes = []
        for entry in changelog:
            try:
                updated_at = datetime.fromisoformat(entry['updatedAt'].replace('Z', '+00:00')).replace(tzinfo=None)
                for field in entry.get('fields', []):
                    if field.get('field', {}).get('id') == 'status':
                        changes.append(TaskChange(
                            timestamp=updated_at,
                            field='status',
                            old_value=field.get('from', {}).get('key', ''),
                            new_value=field.get('to', {}).get('key', '')
                        ))
            except:
                continue
        return sorted(changes, key=lambda x: x.timestamp)
    
    def get_status_on_date(self, task: Task, target_date: datetime) -> Optional[str]:
        if target_date < task.created:
            return None
        if not task.changes:
            return task.status
        current_status = task.status
        sorted_changes = sorted(task.changes, key=lambda x: x.timestamp, reverse=True)
        for change in sorted_changes:
            if change.timestamp > target_date:
                current_status = change.old_value
            else:
                break
        return current_status

    def get_status_dates_in_month(self, task: Task, status: str, year: int, month: int):
        dates = []
        for change in task.changes:
            if (change.field == 'status' and 
                change.new_value == status and
                change.timestamp.year == year and
                change.timestamp.month == month):
                dates.append(change.timestamp)
        return (dates[0] if dates else None, dates[-1] if dates else None)

    def process_tasks_for_month(self, year: int, month: int) -> pd.DataFrame:
        first_day = datetime(year, month, 1)
        if month == 12:
            last_day = datetime(year + 1, 1, 1) - timedelta(days=1)
        else:
            last_day = datetime(year, month + 1, 1) - timedelta(days=1)
        
        all_tasks_data = []
        
        for project_name in TARGET_PROJECTS:
            raw_tasks = self.get_tasks_by_project(project_name, year, month)
            if not raw_tasks:
                continue
                
            logger.info(f"Проект {project_name}: обрабатываю {len(raw_tasks)} задач")
            self.fetch_changelogs_concurrently(raw_tasks)
            
            for raw_task in raw_tasks:
                try:
                    created_at = datetime.fromisoformat(raw_task['createdAt'].replace('Z', '+00:00')).replace(tzinfo=None)
                    updated_at = datetime.fromisoformat(raw_task['updatedAt'].replace('Z', '+00:00')).replace(tzinfo=None)
                    resolved_at = None
                    if raw_task.get('resolvedAt'):
                        resolved_at = datetime.fromisoformat(raw_task['resolvedAt'].replace('Z', '+00:00')).replace(tzinfo=None)

                    # Получаем ключ родителя
                    parent_key = None
                    if 'parent' in raw_task and raw_task['parent']:
                        parent_key = raw_task['parent']['key']

                    task = Task(
                        key=raw_task['key'],
                        summary=raw_task['summary'],
                        project=project_name,
                        assignee=raw_task.get('assignee', {}).get('display', 'Не назначен'),
                        status=raw_task['status']['key'],
                        created=created_at,
                        updated=updated_at,
                        resolved=resolved_at,
                        priority=raw_task.get('priority', {}).get('display', ''),
                        parent_key=parent_key
                    )
                    
                    # Вычисляем раздел и глубину
                    section, depth = self.resolve_root_section(task)
                    
                    changelog = self.get_task_changelog(task.key)
                    task.changes = self.extract_status_changes(task.key, changelog)
                    
                    status_on_first = self.get_status_on_date(task, first_day)
                    
                    status_changed_in_month = any(
                        c.timestamp.year == year and c.timestamp.month == month 
                        for c in task.changes
                    )
                    
                    should_include = False
                    if status_changed_in_month:
                        should_include = True
                    elif status_on_first is not None and status_on_first != TaskStatus.CLOSED:
                        should_include = True
                    elif status_on_first is None:
                        if task.created.year == year and task.created.month == month:
                            should_include = True

                    if not should_include:
                        continue

                    status_columns = {}
                    status_flags = {}
                    
                    for st in TaskStatus.ALL_STATUSES:
                        first_d, last_d = self.get_status_dates_in_month(task, st, year, month)
                        status_columns[f'first_{st}_date'] = first_d if first_d else '-'
                        status_columns[f'last_{st}_date'] = last_d if last_d else '-'
                        was_in_status = (status_on_first == st) or (first_d is not None)
                        status_flags[f'was_{st}_in_month'] = 1 if was_in_status else 0

                    was_in_work = status_flags[f'was_{TaskStatus.IN_PROGRESS}_in_month']
                    closed_in_month = (status_columns[f'last_{TaskStatus.CLOSED}_date'] != '-')
                    
                    row = {
                        'Ключ': task.key,
                        'Ссылка': f"https://tracker.yandex.ru/{task.key}",
                        'Заголовок': task.summary,
                        'Проект': task.project,
                        'Раздел': section,
                        'Уровень вложенности': depth,
                        'Исполнитель': task.assignee,
                        'Текущий статус (API)': TaskStatus.get_display_name(task.status),
                        'Период отчета': f"{month:02d}.{year}",
                        'Приоритет': task.priority,
                        'Дата создания': task.created,
                        'Дата обновления': task.updated,
                        'Дата решения': task.resolved,
                        'Статус на 1 число': TaskStatus.get_display_name(status_on_first) if status_on_first else 'Не создана',
                        'Был в "Открыт"': status_flags[f'was_{TaskStatus.OPEN}_in_month'],
                        'Был в "В работе"': was_in_work,
                        'Был в "Приостановлено"': status_flags[f'was_{TaskStatus.PAUSED}_in_month'],
                        'Был в "Требуется инфо"': status_flags[f'was_{TaskStatus.NEED_INFO}_in_month'],
                        'Закрыта в этом месяце': 1 if closed_in_month else 0,
                        'Первый переход в Открыт': status_columns[f'first_{TaskStatus.OPEN}_date'],
                        'Последний переход в Открыт': status_columns[f'last_{TaskStatus.OPEN}_date'],
                        'Первый переход в В работе': status_columns[f'first_{TaskStatus.IN_PROGRESS}_date'],
                        'Последний переход в В работе': status_columns[f'last_{TaskStatus.IN_PROGRESS}_date'],
                        'Первый переход в Паузу': status_columns[f'first_{TaskStatus.PAUSED}_date'],
                        'Последний переход в Паузу': status_columns[f'last_{TaskStatus.PAUSED}_date'],
                        'Первый переход в Закрыт': status_columns[f'first_{TaskStatus.CLOSED}_date'],
                        'Последний переход в Закрыт': status_columns[f'last_{TaskStatus.CLOSED}_date'],
                        'Первый переход в Инфо': status_columns[f'first_{TaskStatus.NEED_INFO}_date'],
                        'Последний переход в Инфо': status_columns[f'last_{TaskStatus.NEED_INFO}_date'],
                    }
                    all_tasks_data.append(row)
                    
                except Exception as e:
                    logger.error(f"Сбой на задаче {raw_task.get('key')}: {e}")
                    continue

        return pd.DataFrame(all_tasks_data)

    def create_pivot_tables(self, df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
        if df.empty:
            return {}
        logger.info("Формирую сводные таблицы...")
        pivots = {}

        # 1. Анализ "В работе"
        work_analysis = []
        for period in df['Период отчета'].unique():
            p_df = df[df['Период отчета'] == period]
            for proj in p_df['Проект'].unique():
                curr = p_df[p_df['Проект'] == proj]
                start_in_work = len(curr[curr['Статус на 1 число'] == 'В работе'])
                arrived_in_work = len(curr[(curr['Был в "В работе"'] == 1) & (curr['Статус на 1 число'] != 'В работе')])
                total_active = curr['Был в "В работе"'].sum()
                done_from_active = len(curr[(curr['Был в "В работе"'] == 1) & (curr['Закрыта в этом месяце'] == 1)])
                
                work_analysis.append({
                    'Период': period,
                    'Проект': proj,
                    'В работе (на начало)': start_in_work,
                    'Пришло в работу': arrived_in_work,
                    'Всего активных (пул)': total_active,
                    'Выполнено из пула': done_from_active,
                    'Остаток в работе': total_active - done_from_active
                })
        pivots['Анализ_В_Работе'] = pd.DataFrame(work_analysis)

        # 2. Сводная по разделам (теперь Раздел - это имя корневой задачи)
        by_section = []
        for period in df['Период отчета'].unique():
            p_df = df[df['Период отчета'] == period]
            for proj in p_df['Проект'].unique():
                proj_df = p_df[p_df['Проект'] == proj]
                for section in sorted(proj_df['Раздел'].unique()):
                    sect_df = proj_df[proj_df['Раздел'] == section]
                    by_section.append({
                        'Период': period,
                        'Проект': proj,
                        'Раздел': section,
                        'Всего задач': len(sect_df),
                        'Был в работе': sect_df['Был в "В работе"'].sum(),
                        'Закрыто': sect_df['Закрыта в этом месяце'].sum()
                    })
        pivots['Сводная_по_разделам'] = pd.DataFrame(by_section)

        # 3. Статусы на 1 число
        on_first = []
        for period in df['Период отчета'].unique():
            p_df = df[df['Период отчета'] == period]
            for proj in p_df['Проект'].unique():
                proj_df = p_df[p_df['Проект'] == proj]
                counts = proj_df['Статус на 1 число'].value_counts()
                row = {'Период': period, 'Проект': proj, 'Всего': len(proj_df)}
                for st_key in TaskStatus.ALL_STATUSES:
                    disp = TaskStatus.get_display_name(st_key)
                    row[disp] = counts.get(disp, 0)
                on_first.append(row)
        pivots['Статусы_на_1_число'] = pd.DataFrame(on_first)
        return pivots

    def save_excel(self, df: pd.DataFrame, pivots: Dict[str, pd.DataFrame], filename: str):
        logger.info(f"Сохраняю отчет в {filename}")
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Все_Задачи', index=False)
            for name, pivot_df in pivots.items():
                pivot_df.to_excel(writer, sheet_name=name, index=False)
            for sheet in writer.sheets.values():
                for column in sheet.columns:
                    max_length = 0
                    column_letter = get_column_letter(column[0].column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
                        except: pass
                    sheet.column_dimensions[column_letter].width = min(max_length + 2, 60)

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('--start-year', type=int)
    parser.add_argument('--start-month', type=int)
    parser.add_argument('--end-year', type=int)
    parser.add_argument('--end-month', type=int)
    args = parser.parse_args()

    now = datetime.now()
    e_year = args.end_year or now.year
    e_month = args.end_month or now.month
    s_year = args.start_year or e_year
    s_month = args.start_month or e_month

    analyzer = YandexTrackerAnalyzer()
    dfs = []
    curr_y, curr_m = s_year, s_month
    while (curr_y < e_year) or (curr_y == e_year and curr_m <= e_month):
        df = analyzer.process_tasks_for_month(curr_y, curr_m)
        if not df.empty: dfs.append(df)
        curr_m += 1
        if curr_m > 12: curr_m = 1; curr_y += 1
    
    if not dfs:
        logger.warning("Нет данных за указанный период.")
        return

    final_df = pd.concat(dfs, ignore_index=True)
    pivots = analyzer.create_pivot_tables(final_df)
    output_file = f"Tracker_Report_{s_year}_{s_month:02d}-{e_year}_{e_month:02d}.xlsx"
    analyzer.save_excel(final_df, pivots, output_file)
    logger.success("Готово!")

if __name__ == "__main__":
    main()
